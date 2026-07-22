# -*- coding: utf-8 -*-
"""
Read the current ITRS export files into the parquet database.

    python import_detail.py

The bank exports two workbooks, one per direction:
    Raw data/money_transfer_in_detail_*.xlsx   -> Receive  (money into Laos)
    Raw data/money_transfer_out_detail_*.xlsx  -> Payment  (money out of Laos)

They are matched by NAME, not by column position: every column is found from
its Lao (or English) header on row 2, so a file whose columns are reordered,
inserted or renamed still lines up. Add a synonym to HEADER_MAP when the bank
introduces a new wording; nothing else needs to change.

Output: Receive/ITRS_R_<year>.parquet and Payment/ITRS_P_<year>.parquet, in the
exact shape build_dashboard.py already reads, so no other file changes.
"""

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
RAW = ROOT / "Raw data"

# Which workbook is which direction. Globs so a dated filename still matches.
SOURCES = [
    ("money_transfer_in_detail_*.xlsx", "Receive", "P... n/a"),
    ("money_transfer_out_detail_*.xlsx", "Payment", "P... n/a"),
]

# Header text (lower-cased, spaces collapsed) -> parquet column. Both the Lao
# wording seen in the files and a plain English fallback are listed, so an
# English-headed export is read too. Extend freely.
HEADER_MAP = {
    # general
    "ທະນາຄານ": "Banks", "bank": "Banks", "banks": "Banks",
    "ຜູ້ໃຊ້ງານ": "Users", "user": "Users", "users": "Users",
    "ວັນເດືອນປີ ເຮັດທຸລະກຳ": "Date_of_Transaction", "date of transaction": "Date_of_Transaction",
    "ວັນເດືອນປີ ສົ່ງຂໍ້ມູນ": "Date_of_Submission", "date of submission": "Date_of_Submission",
    "ເລກທີ່ໃບອ້າງອີງ": "Reference_Number", "reference number": "Reference_Number",
    # transferor (sender)
    "ລະຫັດຂອງຜູ້ໂອນ": "Transferor_Code", "transferor code": "Transferor_Code",
    "ຊື່ຜູ້ໂອນ": "Transferor_Name", "transferor name": "Transferor_Name",
    "ທີ່ຢູ່ຜູ້ໂອນ": "Transferor_Address", "transferor address": "Transferor_Address",
    "ເບີໂທຕິດຕໍ່ຜູ້ໂອນ": "Transferor_Contact_Number", "transferor contact number": "Transferor_Contact_Number",
    "ເລກວິສະຫະກິດຜູ້ໂອນ": "Transferor_Business_Number", "transferor business number": "Transferor_Business_Number",
    "ເລກປະຈຳຕົວຜູ້ເສຍອາກອນຜູ້ໂອນ": "Transferor_TIN", "transferor tin": "Transferor_TIN",
    "ຊື່ ແລະ ເລກບັນຊີຜູ້ໂອນ": "TransferorName_AccountNumber",
    # recipient (receiver)
    "ລະຫັດຜູ້ຮັບ": "Recipient_Code", "recipient code": "Recipient_Code",
    "ຊື່ຜູ້ຮັບ": "Recipient_Name", "recipient name": "Recipient_Name",
    "ທີ່ຢູ່ຜູ້ຮັບ": "Recipient_Address", "recipient address": "Recipient_Address",
    "ເບີໂທຕິດຕໍ່ຜູ້ຮັບ": "Recipient_Contact_Number", "recipient contact number": "Recipient_Contact_Number",
    "ເລກວິສະຫະກິດຜູ້ຮັບ": "Recipient_Business_Number", "recipient business number": "Recipient_Business_Number",
    "ເລກປະຈຳຕົວຜູ້ເສຍອາກອນຜູ້ຮັບ": "Recipient_TIN", "recipient tin": "Recipient_TIN",
    "ຊື່ ແລະ ເລກບັນຊີຜູ້ຮັບ": "RecipientName_ AccountNumber",
    # documents / relationship
    "ເອກະສານປະກອບ": "Supporting_Documents", "supporting documents": "Supporting_Documents",
    "ເລກທີເອກະສານປະກອບ": "Document_Number", "document number": "Document_Number",
    "ໃບອານຸຍາດຈາກ ທຫລ": "BOL_License", "bol license": "BOL_License",
    "ຄວາມສຳພັນຕໍ່ຜູ້ຮັບ": "Relationship_Recipient", "relationship recipient": "Relationship_Recipient",
    "ລະຫັດ lsic": "LSIC_Code", "lsic": "LSIC_Code", "lsic code": "LSIC_Code",
    # purpose / country / method
    "ລະຫັດຈຸດປະສົງ": "Purpose_Code", "purpose code": "Purpose_Code",
    "ຊື່ຈຸດປະສົງ": "Purpose_Name", "purpose name": "Purpose_Name",
    "ລະຫັດປະເທດ": "Country_Code", "country code": "Country_Code",
    "ວິທີການໂອນ": "Transfer_Method", "transfer method": "Transfer_Method",
    # money
    "ລະຫັດສະກຸນເງິນ": "Currency_Code", "currency code": "Currency_Code",
    "ຈຳນວນເງິນທີ່ໂອນ": "Amount_Transferred", "amount transferred": "Amount_Transferred",
    "ອັດຕາແລກປ່ຽນເປັນກີບ": "Exchange_Rates_Kip", "exchange rate kip": "Exchange_Rates_Kip",
    "ຈຳນວນເງິນເປັນກີບ": "Amount_Kip", "amount kip": "Amount_Kip",
    "ອັດຕາແລກປ່ຽນເປັນໂດລາ": "Exchange_Rates_USD", "exchange rate usd": "Exchange_Rates_USD",
    "ຈຳນວນເງີນເປັນໂດລາ": "Amount_USD", "ຈຳນວນເງິນເປັນໂດລາ": "Amount_USD", "amount usd": "Amount_USD",
    "ຄ່າທຳນຽມການໂອນ": "Transfer_Fee", "transfer fee": "Transfer_Fee",
    "ຂໍ້ມູນເພີ່ມເຕີມ": "Additional_Information", "additional information": "Additional_Information",
}

NUM_COLS = {"Amount_Transferred", "Exchange_Rates_Kip", "Amount_Kip",
            "Exchange_Rates_USD", "Amount_USD", "Transfer_Fee"}
DATE_COLS = {"Date_of_Transaction", "Date_of_Submission"}


def norm_header(v):
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


def find_header_row(ws, limit=6):
    """The header is the first row that resolves several known columns."""
    for r in range(1, limit + 1):
        vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
        hits = sum(1 for v in vals if norm_header(v) in HEADER_MAP)
        if hits >= 6:
            colmap = {}
            for i, v in enumerate(vals):
                fld = HEADER_MAP.get(norm_header(v))
                if fld and fld not in colmap.values():
                    colmap[i] = fld
            return r, colmap
    return None, None


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    # The export prefixes text cells with a lone apostrophe to force text.
    if s.startswith("'"):
        s = s[1:].strip()
    return s or None


def to_num(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except (TypeError, ValueError):
        return None


def load_reference():
    """Method vocabulary and purpose->Move_Fund, taken from 2025 so the new
    rows use the same series names as history."""
    import duckdb
    con = duckdb.connect()
    src = "read_parquet(['%s','%s'], union_by_name=true)" % (
        (ROOT / "Receive" / "ITRS_R_2025.parquet").as_posix(),
        (ROOT / "Payment" / "ITRS_P_2025.parquet").as_posix())
    methods = {}
    for (full,) in con.execute(
            "SELECT DISTINCT Transfer_Method FROM %s WHERE Transfer_Method IS NOT NULL" % src).fetchall():
        head = str(full).split(":")[0].strip().lower()
        methods.setdefault(head, full)
    move = {}
    for pur, mf in con.execute(
            """SELECT Pur_5, MODE(Move_Fund) FROM %s
               WHERE Pur_5 IS NOT NULL AND Move_Fund IS NOT NULL GROUP BY 1""" % src).fetchall():
        move[pur] = mf
    return methods, move


def load_purpose_names():
    p = HERE / "config" / "purpose_names.json"
    return json.loads(p.read_text(encoding="utf-8")).get("names", {}) if p.exists() else {}


def extract(path, flow, purpose_names, methods, move):
    print(f"  {path.name} -> {flow}", flush=True)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hrow, colmap = find_header_row(ws)
    if not colmap:
        sys.exit(f"    could not find a header row in {path.name}; "
                 f"add its column wordings to HEADER_MAP")
    print(f"    header on row {hrow}, {len(colmap)} columns matched")

    out = []
    skipped = 0
    for row_no, raw in enumerate(ws.iter_rows(min_row=hrow + 1, values_only=True), start=hrow + 1):
        d = raw[2] if len(raw) > 2 else None
        if not isinstance(d, datetime):
            skipped += 1
            continue
        bank = clean_text(raw[0])
        if not bank or bank.lower() == "format":
            skipped += 1
            continue

        rec = {}
        for ix, name in colmap.items():
            v = raw[ix] if len(raw) > ix else None
            if name in NUM_COLS:
                rec[name] = to_num(v)
            elif name in DATE_COLS:
                rec[name] = v if isinstance(v, datetime) else None
            else:
                rec[name] = clean_text(v)

        # Derived columns 2017-2025 carry but this export does not.
        code = rec.get("Purpose_Code") or ""
        rec["Code_New_Confirm"] = code                 # the export's code is already the confirmed one
        pur5 = code[:6]
        rec["Pur_5"] = pur5 or None
        rec["Pur_2"] = code[:2] or None
        # Prefer the name the bank filed; fall back to the official Lao name.
        if not rec.get("Purpose_Name"):
            rec["Purpose_Name"] = (purpose_names.get(pur5) or {}).get("lo")
        m = rec.get("Transfer_Method")
        if m:
            rec["Transfer_Method"] = methods.get(m.split(":")[0].strip().lower(), m)
        rec["Move_Fund"] = move.get(pur5, "Yes")
        rec["Flow"] = flow
        rec["Source_Year"] = d.year
        rec["Source_File"] = path.name
        rec["Month_Year"] = datetime(d.year, d.month, 1)
        # Use / M2 are compilation flags the export omits; leave them for the
        # build's rule engine, which recomputes them from the code.
        rec.setdefault("Use", None)
        rec.setdefault("M2", None)
        # Row number keeps genuinely-repeated transfers from colliding.
        rec["Hash_ID"] = hashlib.sha1(
            "|".join([flow, path.name, str(row_no)] + [str(rec.get(k)) for k in
                     ("Banks", "Reference_Number", "Date_of_Transaction",
                      "Amount_Transferred", "Currency_Code", "Purpose_Code")])
            .encode("utf-8")).hexdigest()[:24]
        out.append(rec)

    wb.close()
    print(f"    {len(out):,} rows ({skipped:,} skipped)")
    return out


ALL_FIELDS = [
    "Hash_ID", "Banks", "Users", "Date_of_Transaction", "Date_of_Submission",
    "Reference_Number", "Transferor_Code", "Transferor_Name", "Transferor_Address",
    "Transferor_Contact_Number", "Supporting_Documents", "Document_Number",
    "Transferor_Business_Number", "Transferor_TIN", "TransferorName_AccountNumber",
    "BOL_License", "Relationship_Recipient", "LSIC_Code", "Recipient_Code",
    "Recipient_Name", "Recipient_Address", "Recipient_Contact_Number",
    "Recipient_Business_Number", "Recipient_TIN", "RecipientName_ AccountNumber",
    "Purpose_Code", "Purpose_Name", "Country_Code", "Transfer_Method",
    "Currency_Code", "Amount_Transferred", "Exchange_Rates_Kip", "Amount_Kip",
    "Exchange_Rates_USD", "Amount_USD", "Transfer_Fee", "Additional_Information",
    "Code_New_Confirm", "Pur_2", "Pur_5", "Month_Year", "Move_Fund", "M2", "Use",
    "Flow", "Source_Year", "Source_File",
]


def write_year(rows, flow):
    by_year = defaultdict(list)
    for r in rows:
        by_year[r["Source_Year"]].append(r)
    prefix = "ITRS_R" if flow == "Receive" else "ITRS_P"
    folder = ROOT / ("Receive" if flow == "Receive" else "Payment")
    folder.mkdir(exist_ok=True)
    for year, yr_rows in sorted(by_year.items()):
        cols = {f: [r.get(f) for r in yr_rows] for f in ALL_FIELDS}
        table = pa.table(cols)
        dest = folder / f"{prefix}_{year}.parquet"
        pq.write_table(table, dest)
        print(f"    wrote {dest.name}: {len(yr_rows):,} rows")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="parse and report, write nothing")
    a = ap.parse_args()

    purpose_names = load_purpose_names()
    methods, move = load_reference()
    print(f"  reference: {len(methods)} methods, {len(move)} purpose->Move_Fund")

    for pattern, flow, _ in SOURCES:
        matches = sorted(RAW.glob(pattern))
        if not matches:
            print(f"  (no file matching {pattern})")
            continue
        rows = []
        for path in matches:
            rows += extract(path, flow, purpose_names, methods, move)
        if a.dry_run:
            print(f"  DRY RUN: {len(rows):,} {flow} rows parsed, not written")
            continue
        write_year(rows, flow)


if __name__ == "__main__":
    main()
