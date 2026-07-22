# -*- coding: utf-8 -*-
"""
Extract 2026 ITRS data from the macro workbook into the parquet database.

    python import_2026.py --from 2026-01-01 --to 2026-06-20

Reads the "Recieve" and "Payment" sheets of ITRS_DataBaseV.11.1.xlsm and writes
Receive/ITRS_R_2026.parquet and Payment/ITRS_P_2026.parquet in the same shape as
the 2017-2025 files, so build_dashboard.py picks them up with no other change.

Column positions were read from the sheet itself and cross-checked against the
VBA in calculate.bas, which addresses the same sheets by Excel letter.
"""

import argparse
import hashlib
import json
import sys
from datetime import datetime
from pathlib import Path

import pyarrow as pa
import pyarrow.parquet as pq
from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
ROOT = HERE.parent
BOOK = ROOT / "Raw data" / "ITRS_DataBaseV.11.1.xlsm"

# 0-based column index -> parquet column. THE TWO SHEETS DIFFER IN H-L:
# Recieve carries the recipient block there; Payment carries the transferor
# block and puts the recipient later. Applying one map to both silently files
# phone numbers, account strings and LSIC codes into the wrong columns.
COMMON = {
    0:  "Banks",                       # A
    1:  "Users",                       # B
    2:  "Date_of_Transaction",         # C
    3:  "Date_of_Submission",          # D
    4:  "Reference_Number",            # E
    5:  "Transferor_Code",             # F
    6:  "Transferor_Name",             # G
    12: "Purpose_Code",                # M  raw code as reported
    13: "Country_Code",                # N
    14: "Transfer_Method",             # O
    15: "Currency_Code",               # P
    16: "Amount_Transferred",          # Q
    17: "Exchange_Rates_Kip",          # R
    18: "Amount_Kip",                  # S
    19: "Exchange_Rates_USD",          # T
    21: "Transfer_Fee",                # V
    22: "Additional_Information",      # W
    23: "Code_New_Confirm",            # X  corrected code, full length
    24: "Pur_2",                       # Y
    25: "Amount_USD",                  # Z  "New Amount": what the workbook reports on
    26: "M2",                          # AA
    27: "Use",                         # AB
}
SHEET_COLS = {
    "Recieve": {**COMMON, 
        7:  "Recipient_Code",                # H
        8:  "Recipient_Name",                # I
        9:  "Recipient_Contact_Number",      # J
        10: "RecipientName_ AccountNumber",  # K
        11: "LSIC_Code",                     # L
    },
    "Payment": {**COMMON, 
        7:  "Transferor_Contact_Number",     # H
        8:  "TransferorName_AccountNumber",  # I
        9:  "LSIC_Code",                     # J
        10: "Recipient_Code",                # K
        11: "Recipient_Name",                # L
    },
}

TEXT_COLS = {
    "Banks", "Users", "Reference_Number", "Transferor_Code", "Transferor_Name",
    "Recipient_Code", "Recipient_Name", "Recipient_Contact_Number",
    "RecipientName_ AccountNumber", "LSIC_Code", "Purpose_Code", "Country_Code",
    "Transfer_Method", "Currency_Code", "Additional_Information", "Pur_5",
    "Pur_2", "M2", "Use", "Code_New_Confirm", "Recipient_Code",
    "Transferor_Contact_Number", "TransferorName_AccountNumber",
}
NUM_COLS = {
    "Amount_Transferred", "Exchange_Rates_Kip", "Amount_Kip",
    "Exchange_Rates_USD", "Amount_USD", "Transfer_Fee",
}


def clean_text(v):
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def to_num(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_reference():
    """Vocabularies taken from the 2017-2025 files, so 2026 lines up with them.

    Transfer_Method loses its Lao gloss in 2026 ("Swift" instead of
    "Swift: ..."), which would split every method into two series. Move_Fund
    is not produced by this workbook version at all, but it is a stable
    function of the purpose code, so it can be recovered from history.
    """
    import duckdb
    con = duckdb.connect()
    src = "read_parquet(['%s','%s'], union_by_name=true)" % (
        (ROOT / "Receive" / "ITRS_R_2025.parquet").as_posix(),
        (ROOT / "Payment" / "ITRS_P_2025.parquet").as_posix())
    methods = {}
    for full, in con.execute(
            "SELECT DISTINCT Transfer_Method FROM %s WHERE Transfer_Method IS NOT NULL" % src).fetchall():
        head = str(full).split(":")[0].strip().lower()
        if head and head not in methods:
            methods[head] = full
    move = {}
    for pur, mf in con.execute(
            """SELECT Pur_5, MODE(Move_Fund) FROM %s
               WHERE Pur_5 IS NOT NULL AND Move_Fund IS NOT NULL
               GROUP BY 1""" % src).fetchall():
        move[pur] = mf
    print("    reference: %d transfer methods, %d purpose->Move_Fund entries"
          % (len(methods), len(move)))
    return methods, move


def load_purpose_names():
    path = HERE / "config" / "purpose_names.json"
    if not path.exists():
        return {}
    return json.loads(path.read_text(encoding="utf-8")).get("names", {})


def extract(sheet, flow, d_from, d_to, purpose_names, methods, move):
    print(f"  reading sheet {sheet!r} ...", flush=True)
    wb = load_workbook(BOOK, read_only=True, data_only=True)
    ws = wb[sheet]
    rows, skipped_date, skipped_template = [], 0, 0

    for row_no, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        d = raw[2] if len(raw) > 2 else None
        if not isinstance(d, datetime):
            skipped_date += 1
            continue
        if not (d_from <= d.date() <= d_to):
            continue
        bank = clean_text(raw[0])
        # The sheet carries a template row whose bank cell reads "Format".
        if bank is None or bank.lower() == "format":
            skipped_template += 1
            continue

        cols = SHEET_COLS[sheet]
        rec = {}
        for ix, name in cols.items():
            v = raw[ix] if len(raw) > ix else None
            if name in NUM_COLS:
                rec[name] = to_num(v)
            elif name.startswith("Date_"):
                rec[name] = v if isinstance(v, datetime) else None
            else:
                rec[name] = clean_text(v)

        # 2017-2025 store Pur_5 as the corrected code truncated to 6 chars.
        # Keeping the raw 8/10-char value would break every Pur_5-keyed series
        # at the year boundary.
        code = rec.get("Code_New_Confirm") or ""
        pur5 = code[:6]
        rec["Pur_5"] = pur5 or None
        rec["Pur_2"] = rec.get("Pur_2") or (code[:2] or None)
        rec["Purpose_Name"] = (purpose_names.get(pur5) or {}).get("lo")
        m = rec.get("Transfer_Method")
        if m:
            rec["Transfer_Method"] = methods.get(m.split(":")[0].strip().lower(), m)
        rec["Flow"] = flow
        rec["Source_Year"] = d.year
        rec["Source_File"] = BOOK.name
        rec["Month_Year"] = datetime(d.year, d.month, 1)
        # Not produced by this workbook version; recovered from the historical
        # purpose -> Move_Fund mapping, which is one-to-one in the 2025 data.
        rec["Move_Fund"] = move.get(pur5, "Yes")
        # Stable per-row id so duplicate detection has something to key on.
        # Row number is part of the key: without it, genuinely repeated
        # transfers collide and Hash_ID stops being unique.
        rec["Hash_ID"] = hashlib.sha1(
            "|".join([flow, str(row_no)] + [str(rec.get(k)) for k in
                     ("Banks", "Reference_Number", "Date_of_Transaction",
                      "Amount_Transferred", "Currency_Code", "Purpose_Code")])
            .encode("utf-8")).hexdigest()[:24]
        # Present in the 2017-2025 files, absent from this sheet.
        for missing in ("Transferor_Address", "Transferor_Contact_Number",
                        "Transferor_Business_Number", "Transferor_TIN",
                        "TransferorName_AccountNumber", "Recipient_Address",
                        "Recipient_Business_Number", "Recipient_TIN",
                        "Supporting_Documents", "Document_Number", "BOL_License",
                        "Relationship_Recipient", "Recipient_Contact_Number",
                        "RecipientName_ AccountNumber", "LSIC_Code"):
            rec.setdefault(missing, None)
        rows.append(rec)

    print(f"    {len(rows):,} rows in range "
          f"(skipped {skipped_date:,} undated, {skipped_template:,} template)", flush=True)
    return rows


def write_parquet(rows, path):
    if not rows:
        print(f"    nothing to write for {path.name}")
        return
    cols = {k: [r.get(k) for r in rows] for k in rows[0]}
    arrays, names = [], []
    for k, vals in cols.items():
        if k in NUM_COLS or k == "Source_Year":
            arr = pa.array(vals, type=pa.float64() if k in NUM_COLS else pa.int64())
        elif k.startswith("Date_") or k == "Month_Year":
            arr = pa.array(vals, type=pa.timestamp("ns"))
        else:
            arr = pa.array([None if v is None else str(v) for v in vals], type=pa.string())
        arrays.append(arr)
        names.append(k)
    path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(pa.Table.from_arrays(arrays, names=names), path, compression="snappy")
    print(f"    wrote {path}  ({path.stat().st_size/1048576:.1f} MB)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--from", dest="d_from", default="2026-01-01")
    ap.add_argument("--to", dest="d_to", default="2026-06-20")
    a = ap.parse_args()
    d_from = datetime.strptime(a.d_from, "%Y-%m-%d").date()
    d_to = datetime.strptime(a.d_to, "%Y-%m-%d").date()
    if not BOOK.exists():
        sys.exit(f"Source workbook not found: {BOOK}")

    print(f"Importing {d_from} to {d_to} from {BOOK.name}")
    names = load_purpose_names()
    methods, move = load_reference()
    for sheet, flow, out in (("Recieve", "Receive", ROOT / "Receive" / "ITRS_R_2026.parquet"),
                             ("Payment", "Payment", ROOT / "Payment" / "ITRS_P_2026.parquet")):
        write_parquet(extract(sheet, flow, d_from, d_to, names, methods, move), out)
    print("\nDone. Run build_dashboard.py to pick the new files up.")


if __name__ == "__main__":
    main()
